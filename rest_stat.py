import csv
import openpyxl

rest = []
model = []
dmb = []

# Выбираем из входного csv список ресторанов, список моделей мониторов и общий список [ресторан, модель, количество]
with open('rest.csv', newline='') as csvfile:
    reader = csv.reader(csvfile, delimiter=';')
    for row in reader:
        rest.append(row[0])
        model.append(row[1].strip())
        dmb.append(row)

# Выбираем уникальные значения ресторанов и моделей дмб
rest_list = set(rest)
model_list = set(model)

# Создаем книгу Excel даем название рабочему листу и задаем ширину первой колонки, чтобы влазили названия ресторанов
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'DMB List'
ws.column_dimensions['A'].width = 30

# Заполняем названия строк и колонок таблицы. Рестораны - строки, модели мониторов - колонки.
restlist = list(rest_list)
restlist.sort(reverse=True)
for row in range(2, 2+len(restlist)):
    ws.cell(row=row, column=1, value=restlist.pop())
modellist = list(model_list)
modellist.sort(reverse=True)
for col in range(2, 2+len(modellist)):
    ws.cell(row=1, column=col, value=modellist.pop())

# Функция для заполнения получившейся таблицы. В качестве аргументов берет название ресторана, название модели ДМБ и количество, которые представлены у нас в общем списке dmb
# Координаты ячейки выбираются поиском по колонке с названием ресторана и строке с моделями мониторов.
# Так как данные csv могли быть не очищены полностью, то перед записью значения проверяем пуста ли ячейка и при необходимости суммируем значение.
def fill_table(r, c, d):
    data_row = data_col = data_val = ''
    for row in ws.iter_rows(max_col=1):
        for cell in row:
            if cell.value == r:
                data_row = cell.row
    for column in ws.iter_cols():
        for cell in column:
            if cell.value == c.strip():
                data_col = cell.column
    if ws.cell(row=data_row, column=data_col).value == None:
        data_val = int(d)
    else:
        data_val = int(ws.cell(row=data_row, column=data_col).value) + int(d)
    ws.cell(row=data_row, column=data_col, value=data_val)
    
# Обходим список количества моделей дмб по ресторанам и заполняем таблицу
for x in dmb:
    fill_table(x[0], x[1], x[2])

# Закрепляем первую строку таблицы с названиями моделй, чтобы при прокрутке их было видно и сохраняем файл Excel
ws.freeze_panes = 'A2'
wb.save('models.xlsx')